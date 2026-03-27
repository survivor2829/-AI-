"""
物保云产品详情页生成器 - Web 后端
启动: python app.py
访问: http://localhost:5000
"""
import os
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
import uuid
import zipfile
import json
from pathlib import Path
from flask import Flask, request, jsonify, send_file, send_from_directory
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "output"
TEMPLATES_DIR = BASE_DIR / "templates"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

ALLOWED_IMG = {"jpg", "jpeg", "png", "webp"}
ALLOWED_DOC = {"pdf", "docx"}
ALLOWED_XLS = {"xlsx"}

# ── DeepSeek API 配置（可替换为其他 OpenAI 兼容接口）─────────────────
DEEPSEEK_API_KEY = "***REMOVED***"  # 替换为你的 Key
DEEPSEEK_API_URL = "https://api.deepseek.com/v1/chat/completions"
DEEPSEEK_MODEL   = "deepseek-reasoner"
PROXY = {"http": "http://127.0.0.1:7890", "https": "http://127.0.0.1:7890"}

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100 MB


# ── 工具函数 ─────────────────────────────────────────────────────────

def allowed_img(fn): return "." in fn and fn.rsplit(".", 1)[1].lower() in ALLOWED_IMG
def allowed_doc(fn): return "." in fn and fn.rsplit(".", 1)[1].lower() in ALLOWED_DOC
def allowed_xls(fn): return "." in fn and fn.rsplit(".", 1)[1].lower() in ALLOWED_XLS


def build_config(data: dict) -> dict:
    """将前端 JSON 表单数据映射为 render.py 所需的配置字典。"""
    core_params = {}
    # 先从专用输入框取
    if data.get("param_efficiency"): core_params["工作效率"] = data["param_efficiency"]
    if data.get("param_width"):      core_params["清扫宽度"] = data["param_width"]
    if data.get("param_capacity"):   core_params["尘箱容量"] = data["param_capacity"]
    if data.get("param_runtime"):    core_params["工作时间"] = data["param_runtime"]

    # 如果从专用框拿到的不足4个，从 detail_params 或 core_params 字段直接补
    if len(core_params) < 4:
        # 优先从前端传来的 core_params 字段
        if isinstance(data.get("core_params"), dict):
            for k, v in data["core_params"].items():
                if k not in core_params and len(core_params) < 4:
                    core_params[k] = v
        # 还不足4个就从 detail_params 取
        if len(core_params) < 4 and isinstance(data.get("detail_params"), dict):
            for k, v in data["detail_params"].items():
                if k not in core_params and len(core_params) < 4:
                    core_params[k] = v

    # detail_params：core_params 作为基底，前端传来的字典可覆盖/追加
    if isinstance(data.get("detail_params"), dict) and data["detail_params"]:
        detail_params = {**core_params, **data["detail_params"]}
    else:
        detail_params = dict(core_params)

    # 产品尺寸：如果 detail_params 里没有则自动补入
    dims_tmp = data.get("dimensions", {})
    _l = dims_tmp.get("length", "") or data.get("dim_length", "")
    _w = dims_tmp.get("width",  "") or data.get("dim_width",  "")
    _h = dims_tmp.get("height", "") or data.get("dim_height", "")
    if _l and _w and _h and "产品尺寸" not in detail_params:
        detail_params["产品尺寸"] = f"{_l}*{_w}*{_h}"

    advantages = [a for a in data.get("advantages", []) if a and a.strip()]

    dims = data.get("dimensions", {})
    return {
        "brand":          data.get("brand", ""),
        "brand_en":       data.get("brand_en", ""),
        "model":          data.get("model", "product"),
        "product_name":   data.get("product_name") or data.get("model", ""),
        "product_type":   data.get("product_type", ""),
        "slogan":         data.get("slogan", ""),
        "sub_slogan":     data.get("sub_slogan", ""),
        "product_image":  data.get("product_image", ""),
        "scene_image":    data.get("scene_image", ""),
        "efficiency_claim": data.get("efficiency_claim", ""),
        "efficiency_value": data.get("efficiency_value", ""),
        "savings_claim":    data.get("savings_claim", ""),
        "people_count":     data.get("people_count", "3"),
        "core_params":    core_params,
        "detail_params":  detail_params,
        "template_type":  data.get("template_type", ""),
        "advantages":     advantages if advantages else ["省时省力", "高效清洁"],
        "dimensions": {
            "length": dims.get("length", "") or data.get("dim_length", ""),
            "width":  dims.get("width",  "") or data.get("dim_width",  ""),
            "height": dims.get("height", "") or data.get("dim_height", ""),
        },
    }


# ── 基础路由 ─────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_file(BASE_DIR / "web_ui" / "index.html")


@app.route("/api/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "请求中没有文件字段"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "文件名为空"}), 400
    if not allowed_img(file.filename):
        return jsonify({"error": f"不支持的格式，请上传 {', '.join(ALLOWED_IMG)}"}), 400
    ext = file.filename.rsplit(".", 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    save_path = UPLOAD_DIR / filename
    file.save(str(save_path))
    return jsonify({"path": str(save_path), "filename": filename})


@app.route("/api/generate", methods=["POST"])
def generate():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "请求体不是合法的 JSON"}), 400

    missing = [k for k in ["model", "product_type", "slogan"] if not data.get(k)]
    if missing:
        return jsonify({"error": f"缺少必填字段: {', '.join(missing)}"}), 400

    config = build_config(data)
    dp = config.get('detail_params')
    print(f"[DEBUG] detail_params keys={list(dp.keys()) if isinstance(dp,dict) else repr(dp)}")
    print(f"[DEBUG] template_type={config.get('template_type')!r}")
    try:
        from render import generate_detail_page
        out_path = generate_detail_page(config, scale=data.get("scale", 2),
                                        base_url="http://127.0.0.1:5000")
        filename = Path(out_path).name
        return jsonify({"filename": filename, "url": f"/api/output/{filename}"})
    except Exception as exc:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(exc)}), 500


@app.route("/api/output/<path:filename>")
def serve_output(filename):
    return send_from_directory(str(OUTPUT_DIR), filename)


@app.route("/api/uploads/<path:filename>")
def serve_upload(filename):
    return send_from_directory(str(UPLOAD_DIR), filename)


# ── 模板列表 ─────────────────────────────────────────────────────────

@app.route("/api/templates", methods=["GET"])
def list_templates():
    """扫描 templates/ 目录，返回真实模板列表。"""
    result = []
    if TEMPLATES_DIR.exists():
        for d in sorted(TEMPLATES_DIR.iterdir()):
            if not d.is_dir():
                continue
            thumb = None
            for ext in ("jpg", "jpeg", "png"):
                s2 = d / f"screen2.{ext}"
                if s2.exists():
                    thumb = f"/api/template-thumb/{d.name}/screen2.{ext}"
                    break
            result.append({"name": d.name, "product_type": d.name, "thumb": thumb})
    return jsonify(result)


@app.route("/api/template-thumb/<product_type>/<filename>")
def serve_template_thumb(product_type, filename):
    return send_from_directory(str(TEMPLATES_DIR / product_type), filename)


# ── 文本解析（DeepSeek API）──────────────────────────────────────────

@app.route("/api/parse-text", methods=["POST"])
def parse_text():
    """
    接收原始文本（JSON 字符串或普通文字描述），返回结构化产品数据 JSON。
    - 如果文本本身是合法 JSON，直接返回
    - 否则调用 DeepSeek API 解析
    """
    data = request.get_json(silent=True)
    if not data or not data.get("text"):
        return jsonify({"error": "缺少 text 字段"}), 400

    raw_text = data["text"].strip()
    if not raw_text:
        return jsonify({"error": "文本内容为空"}), 400

    # 先尝试直接 JSON 解析
    try:
        parsed = json.loads(raw_text)
        if isinstance(parsed, dict):
            return jsonify(parsed)
    except (json.JSONDecodeError, ValueError):
        pass

    # 调用 DeepSeek API
    try:
        result = _call_deepseek_parse(raw_text)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"AI 解析失败: {e}"}), 500


def _call_deepseek_parse(raw_text: str) -> dict:
    import requests as req
    resp = req.post(
        DEEPSEEK_API_URL,
        headers={"Authorization": f"Bearer {DEEPSEEK_API_KEY}"},
        json={
            "model": DEEPSEEK_MODEL,
            "messages": [
                {
                    "role": "system",
                    "content": "你是产品数据解析助手，把用户提供的产品信息提取为标准JSON格式"
                },
                {
                    "role": "user",
                    "content": (
                        "请把以下产品数据解析为JSON，字段包括：brand, brand_en, product_name, model, "
                        "product_type（洗地机/扫地车/洗扫机器人/高压清洗机/工业吸尘器之一）, "
                        "slogan, sub_slogan, efficiency_claim, savings_claim, "
                        "param_efficiency（工作效率）, param_width（清扫/清洗宽度）, "
                        "param_capacity（尘箱/水箱容量）, param_runtime（续航/工作时间）, "
                        "detail_params（所有技术参数的完整键值对dict，尽可能提取，"
                        "如清洗宽度/吸水宽度/清水容量/污水容量/工作效率/工作时间/"
                        "刷盘电机/吸水电机/刷盘压力/工作噪音/电瓶容量/整机重量/产品尺寸等）, "
                        "advantages(数组,最多6个), dimensions(length/width/height)。"
                        "只返回JSON，不要其他文字：\n\n" + raw_text
                    )
                }
            ],
            "temperature": 0.1,
        },
        proxies=PROXY,
        timeout=120,
    )
    resp.raise_for_status()
    msg = resp.json()["choices"][0]["message"]
    # deepseek-reasoner 有 reasoning_content 和 content 两个字段，取 content
    raw = (msg.get("content") or "").strip()

    # 从 markdown 代码块里提取 JSON
    if "```" in raw:
        import re as _re
        m = _re.search(r"```(?:json)?\s*([\s\S]+?)```", raw)
        if m:
            raw = m.group(1).strip()

    # 如果整段里有多个 {} 块，取第一个完整 JSON 对象
    if not raw.startswith("{"):
        start = raw.find("{")
        if start != -1:
            raw = raw[start:]

    return json.loads(raw.strip())


# ── Excel 批量解析 ───────────────────────────────────────────────────

COLUMN_MAP = {
    "型号": "model", "产品型号": "model",
    "品牌": "brand", "品牌名称": "brand",
    "英文品牌": "brand_en",
    "产品名称": "product_name", "名称": "product_name",
    "产品类型": "product_type", "类型": "product_type",
    "主标语": "slogan", "标语": "slogan",
    "副标语": "sub_slogan",
    "工作效率": "param_efficiency", "清洁效率": "param_efficiency",
    "清扫宽度": "param_width", "清洁宽度": "param_width",
    "尘箱容量": "param_capacity", "水箱容量": "param_capacity",
    "工作时间": "param_runtime", "续航": "param_runtime",
    "效率对比": "efficiency_claim",
    "节省金额": "savings_claim",
    "优势1": "adv_1", "优势2": "adv_2", "优势3": "adv_3",
    "优势4": "adv_4", "优势5": "adv_5", "优势6": "adv_6",
    "长": "dim_length", "宽": "dim_width", "高": "dim_height",
    "正面图": "img_front", "产品图": "img_front",
    "侧面图": "img_side",
    "细节图1": "img_detail1", "细节图2": "img_detail2",
}


@app.route("/api/parse-excel", methods=["POST"])
def parse_excel():
    """解析 xlsx，返回产品列表（含嵌入图片提取）。"""
    if "file" not in request.files:
        return jsonify({"error": "请求中没有文件字段"}), 400
    file = request.files["file"]
    if not file.filename or not allowed_xls(file.filename):
        return jsonify({"error": "仅支持 xlsx 格式"}), 400

    tmp_path = UPLOAD_DIR / f"excel_{uuid.uuid4().hex}.xlsx"
    file.save(str(tmp_path))

    try:
        products = _parse_xlsx(tmp_path)
        return jsonify({"products": products})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Excel 解析失败: {e}"}), 500
    finally:
        tmp_path.unlink(missing_ok=True)


def _parse_xlsx(path: Path) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col_fields = [COLUMN_MAP.get(h, h) for h in headers]

    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        item = {}
        for i, val in enumerate(row):
            if i < len(col_fields) and col_fields[i]:
                item[col_fields[i]] = str(val).strip() if val is not None else ""

        # 合并 adv_1..6 → advantages list
        advantages = []
        for k in ["adv_1", "adv_2", "adv_3", "adv_4", "adv_5", "adv_6"]:
            v = item.pop(k, "")
            if v:
                advantages.append(v)
        item["advantages"] = advantages
        products.append(item)

    # 尝试提取嵌入图片（best-effort）
    try:
        _extract_xlsx_images(path, products, col_fields)
    except Exception:
        pass

    return products


def _extract_xlsx_images(path: Path, products: list, col_fields: list):
    import openpyxl
    wb2 = openpyxl.load_workbook(str(path), data_only=True)
    ws2 = wb2.active
    for img in getattr(ws2, "_images", []):
        try:
            anchor = img.anchor
            row_idx = anchor._from.row   # 0-based
            col_idx = anchor._from.col   # 0-based
            if row_idx == 0:
                continue
            prod_idx = row_idx - 1
            if prod_idx >= len(products):
                continue
            field = col_fields[col_idx] if col_idx < len(col_fields) else ""
            if field not in ("img_front", "img_side", "img_detail1", "img_detail2"):
                continue
            img_data = img._data()
            filename = f"{uuid.uuid4().hex}.png"
            save_path = UPLOAD_DIR / filename
            save_path.write_bytes(img_data)
            products[prod_idx][field] = str(save_path)
        except Exception:
            continue


# ── 批量生成 ─────────────────────────────────────────────────────────

@app.route("/api/batch-generate", methods=["POST"])
def batch_generate():
    """逐个生成，返回全部结果列表。"""
    data = request.get_json(silent=True)
    if not data or "products" not in data:
        return jsonify({"error": "缺少 products 字段"}), 400

    from render import generate_detail_page
    scale = data.get("scale", 2)
    results = []

    for prod in data["products"]:
        # 图片字段映射
        if prod.get("img_front") and not prod.get("product_image"):
            prod["product_image"] = prod["img_front"]
        if prod.get("img_side") and not prod.get("scene_image"):
            prod["scene_image"] = prod["img_side"]
        try:
            config = build_config(prod)
            out_path = generate_detail_page(config, scale=scale,
                                            base_url="http://127.0.0.1:5000")
            filename = Path(out_path).name
            results.append({
                "model": prod.get("model", "unknown"),
                "success": True,
                "filename": filename,
                "url": f"/api/output/{filename}",
            })
        except Exception as e:
            results.append({
                "model": prod.get("model", "unknown"),
                "success": False,
                "error": str(e),
            })

    return jsonify({"results": results})


# ── 打包下载 ─────────────────────────────────────────────────────────

@app.route("/api/download-zip", methods=["GET"])
def download_zip():
    """打包 output/ 下指定文件为 zip 返回。?files=a.png,b.png"""
    files_param = request.args.get("files", "")
    filenames = [f.strip() for f in files_param.split(",") if f.strip()]
    if not filenames:
        return jsonify({"error": "未指定文件"}), 400

    zip_path = OUTPUT_DIR / f"batch_{uuid.uuid4().hex[:8]}.zip"
    with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
        for fn in filenames:
            fp = OUTPUT_DIR / fn
            if fp.exists():
                zf.write(str(fp), fn)

    return send_file(str(zip_path), as_attachment=True, download_name="batch_output.zip")


# ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 50)
    print("  物保云产品详情页生成器 - Web UI")
    print("=" * 50)
    print("  访问: http://localhost:5000")
    print("=" * 50)
    app.run(debug=True, port=5000, use_reloader=False)
